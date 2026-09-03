import os
import json
from openpyxl import Workbook

PASTA_INSERCAO = "Resultados_insercao"
PASTA_NN = "Resultados_nn"

arquivo = Workbook()
aba = arquivo.active
aba.title = "Resultados"

aba.cell(row=2, column=1, value="Instance")
aba.cell(row=2, column=2, value="Solution")
aba.cell(row=2, column=3, value="Time")

linha = 3

for nome_arquivo in os.listdir(PASTA_NN):
    if nome_arquivo.endswith(".json"):
        caminho = os.path.join(PASTA_NN, nome_arquivo)
        with open(caminho, 'r', encoding="utf-8") as f:
            dados = json.load(f)

        custo = dados["custo"]
        tempo = dados["tempo"]

        nome_instancia = os.path.splitext(nome_arquivo)[0]

        aba.cell(row=linha, column=1, value=nome_instancia)
        aba.cell(row=linha, column=2, value=custo)
        aba.cell(row=linha, column=3, value=tempo)

        linha += 1


aba.cell(row=2, column=4, value="Solution")
aba.cell(row=2, column=5, value="Time")

linha = 3
 
for nome_arquivo in os.listdir(PASTA_INSERCAO):
    if nome_arquivo.endswith(".json"):
        caminho = os.path.join(PASTA_INSERCAO, nome_arquivo)
        with open(caminho, 'r', encoding="utf-8") as f:
            dados = json.load(f)

        custo = dados["custo"]
        tempo = dados["tempo"]

        nome_instancia = os.path.splitext(nome_arquivo)[0]

        aba.cell(row=linha, column=4, value=custo)
        aba.cell(row=linha, column=5, value=tempo)

        linha += 1

arquivo.save("Resultados22072026.xlsx")


